"""
FocusGuard - Report Generation Tests
Tests for PDF and Excel report generation
"""

import pytest
import os
import sys
import tempfile
import shutil

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from server.reports import ReportGenerator


class TestReportGenerator:
    """Test suite for ReportGenerator class"""
    
    @pytest.fixture
    def generator(self):
        """Create a temporary report generator"""
        temp_dir = tempfile.mkdtemp()
        gen = ReportGenerator(output_dir=temp_dir)
        yield gen
        # Cleanup
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    @pytest.fixture
    def sample_data(self):
        """Sample test data"""
        violations = [
            {"timestamp": "2026-02-09 10:15:30", "student_name": "Nguyen Van A", "behavior": "Looking Left", "confidence": 0.85},
            {"timestamp": "2026-02-09 10:18:45", "student_name": "Tran Thi B", "behavior": "Head Down", "confidence": 0.92},
            {"timestamp": "2026-02-09 10:22:10", "student_name": "Nguyen Van A", "behavior": "Talking", "confidence": 0.78},
            {"timestamp": "2026-02-09 11:05:00", "student_name": "Le Van C", "behavior": "Looking Right", "confidence": 0.88},
        ]
        
        participants = [
            {"full_name": "Nguyen Van A", "violation_count": 2, "is_flagged": True},
            {"full_name": "Tran Thi B", "violation_count": 1, "is_flagged": False},
            {"full_name": "Le Van C", "violation_count": 1, "is_flagged": False},
            {"full_name": "Pham Thi D", "violation_count": 0, "is_flagged": False},
        ]
        
        return {
            "exam_name": "Test Exam - Unit Testing",
            "exam_code": "TEST01",
            "exam_date": "2026-02-09",
            "violations": violations,
            "participants": participants
        }
    
    def test_generator_init(self, generator):
        """Test generator initialization"""
        assert generator is not None
        assert os.path.exists(generator.output_dir)
    
    def test_pdf_generation(self, generator, sample_data):
        """Test PDF report generation"""
        pdf_path = generator.generate_pdf_report(
            exam_name=sample_data["exam_name"],
            exam_code=sample_data["exam_code"],
            exam_date=sample_data["exam_date"],
            violations=sample_data["violations"],
            participants=sample_data["participants"]
        )
        
        assert os.path.exists(pdf_path)
        assert pdf_path.endswith(".pdf")
        assert os.path.getsize(pdf_path) > 0
        print(f"✅ PDF generated: {pdf_path} ({os.path.getsize(pdf_path)} bytes)")
    
    def test_excel_generation(self, generator, sample_data):
        """Test Excel report generation"""
        excel_path = generator.generate_excel_report(
            exam_name=sample_data["exam_name"],
            exam_code=sample_data["exam_code"],
            exam_date=sample_data["exam_date"],
            violations=sample_data["violations"],
            participants=sample_data["participants"]
        )
        
        assert os.path.exists(excel_path)
        assert excel_path.endswith(".xlsx")
        assert os.path.getsize(excel_path) > 0
        print(f"✅ Excel generated: {excel_path} ({os.path.getsize(excel_path)} bytes)")
    
    def test_excel_sheets(self, generator, sample_data):
        """Test Excel file has correct sheets"""
        from openpyxl import load_workbook
        
        excel_path = generator.generate_excel_report(
            exam_name=sample_data["exam_name"],
            exam_code=sample_data["exam_code"],
            exam_date=sample_data["exam_date"],
            violations=sample_data["violations"],
            participants=sample_data["participants"]
        )
        
        wb = load_workbook(excel_path)
        sheet_names = wb.sheetnames
        
        assert "Summary" in sheet_names
        assert "Participants" in sheet_names
        assert "Violations" in sheet_names
        print(f"✅ Excel has sheets: {sheet_names}")
    
    def test_statistics(self, generator, sample_data):
        """Test statistics calculation"""
        stats = generator.get_statistics(
            violations=sample_data["violations"],
            participants=sample_data["participants"]
        )
        
        assert stats["total_violations"] == 4
        assert stats["total_participants"] == 4
        assert stats["flagged_participants"] == 1
        assert "violations_by_type" in stats
        assert "violations_by_hour" in stats
        
        # Check violations by type
        by_type = stats["violations_by_type"]
        assert by_type.get("Looking Left") == 1
        assert by_type.get("Head Down") == 1
        assert by_type.get("Talking") == 1
        assert by_type.get("Looking Right") == 1
        
        # Check violations by hour
        by_hour = stats["violations_by_hour"]
        assert by_hour.get("10") == 3
        assert by_hour.get("11") == 1
        
        print(f"✅ Statistics: {stats}")
    
    def test_empty_data(self, generator):
        """Test with empty violations and participants"""
        pdf_path = generator.generate_pdf_report(
            exam_name="Empty Test",
            exam_code="EMPTY1",
            exam_date="2026-02-09",
            violations=[],
            participants=[]
        )
        
        assert os.path.exists(pdf_path)
        print("✅ Empty data handling works")
    
    def test_custom_filename(self, generator, sample_data):
        """Test custom output filename"""
        custom_name = "custom_report_test.pdf"
        pdf_path = generator.generate_pdf_report(
            exam_name=sample_data["exam_name"],
            exam_code=sample_data["exam_code"],
            exam_date=sample_data["exam_date"],
            violations=sample_data["violations"],
            participants=sample_data["participants"],
            output_filename=custom_name
        )
        
        assert pdf_path.endswith(custom_name)
        print(f"✅ Custom filename: {pdf_path}")


class TestStatisticsCalculation:
    """Test suite for statistics calculation edge cases"""
    
    @pytest.fixture
    def generator(self):
        temp_dir = tempfile.mkdtemp()
        gen = ReportGenerator(output_dir=temp_dir)
        yield gen
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    def test_avg_violations_per_student(self, generator):
        """Test average violations per student calculation"""
        violations = [{"behavior": "Test"}] * 10
        participants = [{"is_flagged": False}] * 5
        
        stats = generator.get_statistics(violations, participants)
        assert stats["avg_violations_per_student"] == 2.0
    
    def test_no_participants(self, generator):
        """Test with no participants (avoid division by zero)"""
        violations = [{"behavior": "Test"}]
        participants = []
        
        stats = generator.get_statistics(violations, participants)
        assert stats["avg_violations_per_student"] == 1.0  # 1 / max(0, 1)
    
    def test_all_flagged(self, generator):
        """Test when all participants are flagged"""
        violations = []
        participants = [
            {"is_flagged": True},
            {"is_flagged": True},
            {"is_flagged": True}
        ]
        
        stats = generator.get_statistics(violations, participants)
        assert stats["flagged_participants"] == 3


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
