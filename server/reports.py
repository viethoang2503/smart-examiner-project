"""
FocusGuard Report Generator
Generates PDF and Excel reports for violation data
"""

import os
import sys
from datetime import datetime
from typing import List, Dict, Optional

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add project path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


class ReportGenerator:
    """Generate PDF and Excel reports for exam violations"""
    
    def __init__(self, output_dir: str = None):
        """
        Initialize report generator
        
        Args:
            output_dir: Directory to save reports (default: server/reports)
        """
        if output_dir is None:
            output_dir = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                'reports'
            )
        
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generate_pdf_report(
        self,
        exam_name: str,
        exam_code: str,
        exam_date: str,
        violations: List[Dict],
        participants: List[Dict],
        output_filename: str = None
    ) -> str:
        """
        Generate PDF violation report
        
        Args:
            exam_name: Name of the exam
            exam_code: Exam code
            exam_date: Date of exam
            violations: List of violation records
            participants: List of participant data
            output_filename: Output PDF filename
            
        Returns:
            Path to generated PDF file
        """
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"report_{exam_code}_{timestamp}.pdf"
        
        filepath = os.path.join(self.output_dir, output_filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,  # Center
            spaceAfter=20
        )
        elements.append(Paragraph("BÁO CÁO VI PHẠM BÀI THI", title_style))
        elements.append(Paragraph("EXAM VIOLATION REPORT", title_style))
        elements.append(Spacer(1, 20))
        
        # Exam info
        info_style = styles['Normal']
        elements.append(Paragraph(f"<b>Tên bài thi / Exam Name:</b> {exam_name}", info_style))
        elements.append(Paragraph(f"<b>Mã bài thi / Exam Code:</b> {exam_code}", info_style))
        elements.append(Paragraph(f"<b>Ngày thi / Exam Date:</b> {exam_date}", info_style))
        elements.append(Paragraph(f"<b>Tổng số thí sinh / Total Participants:</b> {len(participants)}", info_style))
        elements.append(Paragraph(f"<b>Tổng số vi phạm / Total Violations:</b> {len(violations)}", info_style))
        elements.append(Spacer(1, 20))
        
        # Statistics section
        elements.append(Paragraph("<b>THỐNG KÊ VI PHẠM / VIOLATION STATISTICS</b>", styles['Heading2']))
        
        # Count violations by type
        violation_counts = {}
        for v in violations:
            behavior = v.get('behavior', 'Unknown')
            violation_counts[behavior] = violation_counts.get(behavior, 0) + 1
        
        if violation_counts:
            stats_data = [['Loại vi phạm / Violation Type', 'Số lượng / Count']]
            for behavior, count in sorted(violation_counts.items(), key=lambda x: -x[1]):
                stats_data.append([behavior, str(count)])
            
            stats_table = Table(stats_data, colWidths=[10*cm, 4*cm])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a4a7a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0f0f0')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(stats_table)
        
        elements.append(Spacer(1, 20))
        
        # Participant summary
        elements.append(Paragraph("<b>DANH SÁCH THÍ SINH / PARTICIPANT LIST</b>", styles['Heading2']))
        
        if participants:
            participant_data = [['STT', 'Tên / Name', 'Vi phạm / Violations', 'Trạng thái / Status']]
            for i, p in enumerate(participants, 1):
                status = "🚩 Flagged" if p.get('is_flagged', False) else "✓ Normal"
                participant_data.append([
                    str(i),
                    p.get('full_name', 'N/A'),
                    str(p.get('violation_count', 0)),
                    status
                ])
            
            participant_table = Table(participant_data, colWidths=[1.5*cm, 6*cm, 3*cm, 4*cm])
            participant_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a4a7a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ]))
            elements.append(participant_table)
        
        elements.append(Spacer(1, 20))
        
        # Violation details
        elements.append(Paragraph("<b>CHI TIẾT VI PHẠM / VIOLATION DETAILS</b>", styles['Heading2']))
        
        if violations:
            violation_data = [['Thời gian / Time', 'Thí sinh / Student', 'Hành vi / Behavior', 'Độ tin cậy / Confidence']]
            for v in violations[:50]:  # Limit to 50 violations in PDF
                violation_data.append([
                    v.get('timestamp', 'N/A')[:19],
                    v.get('student_name', 'N/A'),
                    v.get('behavior', 'N/A'),
                    f"{v.get('confidence', 0)*100:.1f}%"
                ])
            
            violation_table = Table(violation_data, colWidths=[4*cm, 4*cm, 4*cm, 3*cm])
            violation_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c0392b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(violation_table)
        
        # Footer
        elements.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.gray,
            alignment=1
        )
        elements.append(Paragraph(
            f"Generated by FocusGuard AI Proctoring System - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            footer_style
        ))
        
        # Build PDF
        doc.build(elements)
        
        return filepath
    
    def generate_excel_report(
        self,
        exam_name: str,
        exam_code: str,
        exam_date: str,
        violations: List[Dict],
        participants: List[Dict],
        output_filename: str = None
    ) -> str:
        """
        Generate Excel violation report
        
        Args:
            exam_name: Name of the exam
            exam_code: Exam code
            exam_date: Date of exam
            violations: List of violation records
            participants: List of participant data
            output_filename: Output Excel filename
            
        Returns:
            Path to generated Excel file
        """
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"report_{exam_code}_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, output_filename)
        
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a4a7a", end_color="1a4a7a", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === Sheet 1: Summary ===
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        ws_summary['A1'] = "BÁO CÁO TỔNG HỢP / SUMMARY REPORT"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')
        
        summary_data = [
            ["Tên bài thi / Exam Name", exam_name],
            ["Mã bài thi / Exam Code", exam_code],
            ["Ngày thi / Exam Date", exam_date],
            ["Tổng số thí sinh / Total Participants", len(participants)],
            ["Tổng số vi phạm / Total Violations", len(violations)],
        ]
        
        for i, row in enumerate(summary_data, 3):
            ws_summary[f'A{i}'] = row[0]
            ws_summary[f'B{i}'] = row[1]
            ws_summary[f'A{i}'].font = Font(bold=True)
        
        # Violation statistics
        ws_summary['A10'] = "THỐNG KÊ VI PHẠM / VIOLATION STATISTICS"
        ws_summary['A10'].font = Font(bold=True, size=12)
        
        violation_counts = {}
        for v in violations:
            behavior = v.get('behavior', 'Unknown')
            violation_counts[behavior] = violation_counts.get(behavior, 0) + 1
        
        ws_summary['A11'] = "Loại vi phạm / Type"
        ws_summary['B11'] = "Số lượng / Count"
        ws_summary['A11'].font = header_font
        ws_summary['B11'].font = header_font
        ws_summary['A11'].fill = header_fill
        ws_summary['B11'].fill = header_fill
        
        for i, (behavior, count) in enumerate(sorted(violation_counts.items(), key=lambda x: -x[1]), 12):
            ws_summary[f'A{i}'] = behavior
            ws_summary[f'B{i}'] = count
        
        # === Sheet 2: Participants ===
        ws_participants = wb.create_sheet("Participants")
        
        participant_headers = ["STT / No.", "Tên / Name", "Vi phạm / Violations", "Đánh dấu / Flagged"]
        for col, header in enumerate(participant_headers, 1):
            cell = ws_participants.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        for i, p in enumerate(participants, 1):
            ws_participants.cell(row=i+1, column=1, value=i)
            ws_participants.cell(row=i+1, column=2, value=p.get('full_name', 'N/A'))
            ws_participants.cell(row=i+1, column=3, value=p.get('violation_count', 0))
            ws_participants.cell(row=i+1, column=4, value="Yes" if p.get('is_flagged', False) else "No")
        
        # === Sheet 3: Violations ===
        ws_violations = wb.create_sheet("Violations")
        
        violation_headers = ["Thời gian / Time", "Thí sinh / Student", "Hành vi / Behavior", "Độ tin cậy / Confidence"]
        for col, header in enumerate(violation_headers, 1):
            cell = ws_violations.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="c0392b", end_color="c0392b", fill_type="solid")
            cell.alignment = center_align
            cell.border = thin_border
        
        for i, v in enumerate(violations, 1):
            ws_violations.cell(row=i+1, column=1, value=v.get('timestamp', 'N/A'))
            ws_violations.cell(row=i+1, column=2, value=v.get('student_name', 'N/A'))
            ws_violations.cell(row=i+1, column=3, value=v.get('behavior', 'N/A'))
            ws_violations.cell(row=i+1, column=4, value=f"{v.get('confidence', 0)*100:.1f}%")
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_participants, ws_violations]:
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save workbook
        wb.save(filepath)
        
        return filepath
    
    def get_statistics(self, violations: List[Dict], participants: List[Dict]) -> Dict:
        """
        Calculate violation statistics
        
        Args:
            violations: List of violation records
            participants: List of participant data
            
        Returns:
            Dictionary containing statistics
        """
        # Count by type
        by_type = {}
        for v in violations:
            behavior = v.get('behavior', 'Unknown')
            by_type[behavior] = by_type.get(behavior, 0) + 1
        
        # Count by hour
        by_hour = {}
        for v in violations:
            timestamp = v.get('timestamp', '')
            if len(timestamp) >= 13:
                hour = timestamp[11:13]
                by_hour[hour] = by_hour.get(hour, 0) + 1
        
        # Flagged count
        flagged_count = sum(1 for p in participants if p.get('is_flagged', False))
        
        return {
            "total_violations": len(violations),
            "total_participants": len(participants),
            "flagged_participants": flagged_count,
            "violations_by_type": by_type,
            "violations_by_hour": by_hour,
            "avg_violations_per_student": len(violations) / max(len(participants), 1)
        }


# Standalone test
if __name__ == "__main__":
    generator = ReportGenerator()
    
    # Sample data
    violations = [
        {"timestamp": "2026-02-07 10:15:30", "student_name": "Nguyen Van A", "behavior": "Looking Left", "confidence": 0.85},
        {"timestamp": "2026-02-07 10:18:45", "student_name": "Tran Thi B", "behavior": "Head Down", "confidence": 0.92},
        {"timestamp": "2026-02-07 10:22:10", "student_name": "Nguyen Van A", "behavior": "Talking", "confidence": 0.78},
    ]
    
    participants = [
        {"full_name": "Nguyen Van A", "violation_count": 2, "is_flagged": False},
        {"full_name": "Tran Thi B", "violation_count": 1, "is_flagged": False},
        {"full_name": "Le Van C", "violation_count": 0, "is_flagged": False},
    ]
    
    # Generate reports
    pdf_path = generator.generate_pdf_report(
        exam_name="Kiểm tra giữa kỳ Toán",
        exam_code="ABC123",
        exam_date="2026-02-07",
        violations=violations,
        participants=participants
    )
    print(f"✅ PDF Report: {pdf_path}")
    
    excel_path = generator.generate_excel_report(
        exam_name="Kiểm tra giữa kỳ Toán",
        exam_code="ABC123",
        exam_date="2026-02-07",
        violations=violations,
        participants=participants
    )
    print(f"✅ Excel Report: {excel_path}")
    
    # Statistics
    stats = generator.get_statistics(violations, participants)
    print(f"\n📊 Statistics: {stats}")
